# attendence/views.py

import secrets
import csv
import math
from io import BytesIO
from datetime import datetime, timedelta, time, date

from django.http import HttpResponse
from django.utils.timezone import localdate, localtime

from django.db.models import Q

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status, permissions
from rest_framework.permissions import IsAuthenticated, IsAdminUser

from rest_framework_simplejwt.views import TokenObtainPairView
from rest_framework_simplejwt.serializers import TokenObtainPairSerializer

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet

from .serializers import (
    RegisterSerializer,
    VerifyOtpSerializer,
    ResendOtpSerializer,
    MeSerializer,

    AttendanceMarkSerializer,
    AttendanceListSerializer,

    OfficeLocationSerializer,
    OfficeQRSerializer,

    LeaveRequestSerializer,
    AdminLeaveDecisionSerializer,
    RegularizationRequestSerializer,
    AdminRegularizationDecisionSerializer,
    ResignationRequestSerializer,
    AdminResignationDecisionSerializer,

    EmployeeDocumentSerializer,
    ESICProfileSerializer,

    OfflineAttendanceRequestSerializer,
    AdminOfflineDecisionSerializer,

    RosterShiftSerializer,
    RosterAssignmentSerializer,

    AdminUserListSerializer,

    DailyReportSerializer,
)

from .models import (
    User,
    Attendance,
    OfficeLocation,
    OfficeQR,

    LeaveRequest,
    RegularizationRequest,
    ResignationRequest,

    EmployeeDocument,
    ESICProfile,
    OfflineAttendanceRequest,

    RosterShift,
    RosterAssignment,

    DailyReport,
)


# ============================================================
# AUTH VIEWS
# ============================================================
class RegisterView(APIView):
    permission_classes = [permissions.AllowAny]

    def post(self, request):
        ser = RegisterSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        ser.save()
        return Response(
            {"detail": "Registered successfully. OTP has been sent to your email."},
            status=status.HTTP_201_CREATED
        )


class VerifyOtpView(APIView):
    permission_classes = [permissions.AllowAny]

    def post(self, request):
        ser = VerifyOtpSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        return Response({"detail": "OTP verified successfully. You can login now."}, status=status.HTTP_200_OK)


class ResendOtpView(APIView):
    permission_classes = [permissions.AllowAny]

    def post(self, request):
        ser = ResendOtpSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        return Response({"detail": "OTP resent successfully."}, status=status.HTTP_200_OK)


class CustomTokenObtainPairSerializer(TokenObtainPairSerializer):
    """
    Email + password login
    """
    @classmethod
    def get_token(cls, user):
        token = super().get_token(user)
        token["email"] = user.email
        token["full_name"] = user.full_name or ""
        token["is_staff"] = bool(getattr(user, "is_staff", False))
        token["is_superuser"] = bool(getattr(user, "is_superuser", False))
        return token

    def validate(self, attrs):
        data = super().validate(attrs)
        user = self.user

        if not user.is_verified:
            raise permissions.PermissionDenied("Email not verified. Please verify OTP first.")

        if not user.is_active:
            raise permissions.PermissionDenied("Account inactive. Please contact admin.")

        return data


class LoginView(TokenObtainPairView):
    permission_classes = [permissions.AllowAny]
    serializer_class = CustomTokenObtainPairSerializer


class MeView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        return Response(MeSerializer(request.user).data, status=status.HTTP_200_OK)


# ============================================================
# ATTENDANCE VIEWS
# ============================================================
class AttendanceMarkView(APIView):
    def post(self, request):
        ser = AttendanceMarkSerializer(data=request.data, context={"request": request})
        ser.is_valid(raise_exception=True)
        result = ser.save()
        return Response(result, status=status.HTTP_200_OK)


class MyAttendanceListView(APIView):
    def get(self, request):
        qs = Attendance.objects.filter(user=request.user).select_related("office").order_by("-date")

        from_date = request.query_params.get("from")
        to_date = request.query_params.get("to")

        if from_date:
            qs = qs.filter(date__gte=from_date)
        if to_date:
            qs = qs.filter(date__lte=to_date)

        if not from_date and not to_date:
            qs = qs[:30]

        data = AttendanceListSerializer(qs, many=True).data
        return Response(data, status=status.HTTP_200_OK)


class TodayAttendanceStatusView(APIView):
    def get(self, request):
        today = localdate()
        att = Attendance.objects.filter(user=request.user, date=today).select_related("office").first()
        if not att:
            return Response({"date": str(today), "checked_in": False, "checked_out": False}, status=status.HTTP_200_OK)

        return Response({
            "date": str(today),
            "office": att.office.name if att.office_id else "",
            "checked_in": bool(att.check_in_time),
            "checked_out": bool(att.check_out_time),
            "check_in_time": att.check_in_time,
            "check_out_time": att.check_out_time,
        }, status=status.HTTP_200_OK)


# ============================================================
# ADMIN OFFICE (OFFICE + QR)
# ============================================================
class AdminOfficeListCreateView(APIView):
    permission_classes = [IsAdminUser]

    def get(self, request):
        offices = OfficeLocation.objects.all().order_by("-id")
        return Response(OfficeLocationSerializer(offices, many=True).data)

    def post(self, request):
        ser = OfficeLocationSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        office = ser.save()
        return Response(OfficeLocationSerializer(office).data, status=status.HTTP_201_CREATED)


class AdminOfficeUpdateView(APIView):
    permission_classes = [IsAdminUser]

    def patch(self, request, office_id):
        office = OfficeLocation.objects.filter(id=office_id).first()
        if not office:
            return Response({"detail": "Office not found"}, status=404)

        ser = OfficeLocationSerializer(office, data=request.data, partial=True)
        ser.is_valid(raise_exception=True)
        office = ser.save()
        return Response(OfficeLocationSerializer(office).data, status=200)


class AdminGenerateOfficeQRView(APIView):
    permission_classes = [IsAdminUser]

    def post(self, request, office_id):
        office = OfficeLocation.objects.filter(id=office_id, is_active=True).first()
        if not office:
            return Response({"detail": "Office not found or inactive"}, status=404)

        token = secrets.token_urlsafe(24)
        qr, created = OfficeQR.objects.get_or_create(
            office=office,
            defaults={"qr_token": token, "is_active": True},
        )
        if not created:
            qr.qr_token = token
            qr.is_active = True
            qr.save()

        return Response(OfficeQRSerializer(qr).data, status=200)


class AdminGetOfficeQRView(APIView):
    permission_classes = [IsAdminUser]

    def get(self, request, office_id):
        qr = OfficeQR.objects.select_related("office").filter(office_id=office_id).first()
        if not qr:
            return Response({"detail": "QR not generated yet"}, status=404)
        return Response(OfficeQRSerializer(qr).data, status=200)


# ============================================================
# LEAVES
# ============================================================
class MyLeaveListCreateView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        qs = LeaveRequest.objects.filter(user=request.user).order_by("-created_at")[:100]
        return Response(LeaveRequestSerializer(qs, many=True).data)

    def post(self, request):
        ser = LeaveRequestSerializer(data=request.data, context={"request": request})
        ser.is_valid(raise_exception=True)
        obj = ser.save()
        return Response(LeaveRequestSerializer(obj).data, status=201)


class AdminLeaveListView(APIView):
    permission_classes = [IsAdminUser]

    def get(self, request):
        status_q = request.query_params.get("status")
        qs = LeaveRequest.objects.all().order_by("-created_at")
        if status_q:
            qs = qs.filter(status=status_q)
        return Response(LeaveRequestSerializer(qs, many=True).data)


class AdminLeaveDecideView(APIView):
    permission_classes = [IsAdminUser]

    def post(self, request, leave_id: int):
        obj = LeaveRequest.objects.filter(id=leave_id).first()
        if not obj:
            return Response({"detail": "Not found"}, status=404)
        ser = AdminLeaveDecisionSerializer(data=request.data, context={"request": request})
        ser.is_valid(raise_exception=True)
        ser.update(obj, ser.validated_data)
        return Response(LeaveRequestSerializer(obj).data, status=200)


# ============================================================
# REGULARIZATION
# ============================================================
class MyRegularizationListCreateView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        qs = RegularizationRequest.objects.filter(user=request.user).order_by("-created_at")[:100]
        return Response(RegularizationRequestSerializer(qs, many=True).data)

    def post(self, request):
        ser = RegularizationRequestSerializer(data=request.data, context={"request": request})
        ser.is_valid(raise_exception=True)
        obj = ser.save()
        return Response(RegularizationRequestSerializer(obj).data, status=201)


class AdminRegularizationListView(APIView):
    permission_classes = [IsAdminUser]

    def get(self, request):
        status_q = request.query_params.get("status")
        qs = RegularizationRequest.objects.all().order_by("-created_at")
        if status_q:
            qs = qs.filter(status=status_q)
        return Response(RegularizationRequestSerializer(qs, many=True).data)


class AdminRegularizationDecideView(APIView):
    permission_classes = [IsAdminUser]

    def post(self, request, req_id: int):
        obj = RegularizationRequest.objects.filter(id=req_id).first()
        if not obj:
            return Response({"detail": "Not found"}, status=404)
        ser = AdminRegularizationDecisionSerializer(data=request.data, context={"request": request})
        ser.is_valid(raise_exception=True)
        ser.update(obj, ser.validated_data)
        return Response(RegularizationRequestSerializer(obj).data, status=200)


# ============================================================
# RESIGNATION
# ============================================================
class MyResignationListCreateView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        qs = ResignationRequest.objects.filter(user=request.user).order_by("-created_at")[:100]
        return Response(ResignationRequestSerializer(qs, many=True).data)

    def post(self, request):
        ser = ResignationRequestSerializer(data=request.data, context={"request": request})
        ser.is_valid(raise_exception=True)
        obj = ser.save()
        return Response(ResignationRequestSerializer(obj).data, status=201)


class AdminResignationListView(APIView):
    permission_classes = [IsAdminUser]

    def get(self, request):
        status_q = request.query_params.get("status")
        qs = ResignationRequest.objects.all().order_by("-created_at")
        if status_q:
            qs = qs.filter(status=status_q)
        return Response(ResignationRequestSerializer(qs, many=True).data)


class AdminResignationDecideView(APIView):
    permission_classes = [IsAdminUser]

    def post(self, request, req_id: int):
        obj = ResignationRequest.objects.filter(id=req_id).first()
        if not obj:
            return Response({"detail": "Not found"}, status=404)
        ser = AdminResignationDecisionSerializer(data=request.data, context={"request": request})
        ser.is_valid(raise_exception=True)
        ser.update(obj, ser.validated_data)
        return Response(ResignationRequestSerializer(obj).data, status=200)


# ============================================================
# DOCUMENTS
# ============================================================
class MyDocumentListCreateView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        qs = EmployeeDocument.objects.filter(user=request.user).order_by("-uploaded_at")[:200]
        return Response(EmployeeDocumentSerializer(qs, many=True, context={"request": request}).data)

    def post(self, request):
        ser = EmployeeDocumentSerializer(data=request.data, context={"request": request})
        ser.is_valid(raise_exception=True)
        obj = ser.save()
        return Response(EmployeeDocumentSerializer(obj, context={"request": request}).data, status=201)


class MyDocumentDeleteView(APIView):
    permission_classes = [IsAuthenticated]

    def delete(self, request, doc_id: int):
        obj = EmployeeDocument.objects.filter(id=doc_id, user=request.user).first()
        if not obj:
            return Response({"detail": "Not found"}, status=404)
        obj.file.delete(save=False)
        obj.delete()
        return Response({"detail": "Deleted"}, status=200)


# ============================================================
# ESIC
# ============================================================
class MyESICView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        obj, _ = ESICProfile.objects.get_or_create(user=request.user)
        return Response(ESICProfileSerializer(obj).data, status=200)

    def patch(self, request):
        obj, _ = ESICProfile.objects.get_or_create(user=request.user)
        ser = ESICProfileSerializer(obj, data=request.data, partial=True)
        ser.is_valid(raise_exception=True)
        ser.save()
        return Response(ESICProfileSerializer(obj).data, status=200)


# ============================================================
# OFFLINE ATTENDANCE REQUEST
# ============================================================
class MyOfflineAttendanceListCreateView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        qs = OfflineAttendanceRequest.objects.filter(user=request.user).select_related("office").order_by("-created_at")[:100]
        return Response(OfflineAttendanceRequestSerializer(qs, many=True).data)

    def post(self, request):
        ser = OfflineAttendanceRequestSerializer(data=request.data, context={"request": request})
        ser.is_valid(raise_exception=True)
        obj = ser.save()
        return Response(OfflineAttendanceRequestSerializer(obj).data, status=201)


class AdminOfflineAttendanceListView(APIView):
    permission_classes = [IsAdminUser]

    def get(self, request):
        status_q = request.query_params.get("status")
        qs = OfflineAttendanceRequest.objects.select_related("office", "user").order_by("-created_at")
        if status_q:
            qs = qs.filter(status=status_q)
        return Response(OfflineAttendanceRequestSerializer(qs, many=True).data)


class AdminOfflineAttendanceDecideView(APIView):
    permission_classes = [IsAdminUser]

    def post(self, request, req_id: int):
        obj = OfflineAttendanceRequest.objects.filter(id=req_id).first()
        if not obj:
            return Response({"detail": "Not found"}, status=404)
        ser = AdminOfflineDecisionSerializer(data=request.data, context={"request": request})
        ser.is_valid(raise_exception=True)
        ser.update(obj, ser.validated_data)
        return Response(OfflineAttendanceRequestSerializer(obj).data, status=200)


# ============================================================
# ROSTER
# ============================================================
class MyRosterView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        qs = RosterAssignment.objects.filter(user=request.user).select_related("shift", "office").order_by("-date")

        from_date = request.query_params.get("from")
        to_date = request.query_params.get("to")
        if from_date:
            qs = qs.filter(date__gte=from_date)
        if to_date:
            qs = qs.filter(date__lte=to_date)

        if not from_date and not to_date:
            qs = qs[:30]

        return Response(RosterAssignmentSerializer(qs, many=True).data, status=200)


class AdminShiftListCreateView(APIView):
    permission_classes = [IsAdminUser]

    def get(self, request):
        qs = RosterShift.objects.all().order_by("name")
        return Response(RosterShiftSerializer(qs, many=True).data)

    def post(self, request):
        ser = RosterShiftSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        obj = ser.save()
        return Response(RosterShiftSerializer(obj).data, status=201)


class AdminRosterAssignView(APIView):
    permission_classes = [IsAdminUser]

    def post(self, request):
        ser = RosterAssignmentSerializer(data=request.data)
        ser.is_valid(raise_exception=True)

        data = ser.validated_data
        obj, _ = RosterAssignment.objects.update_or_create(
            user=data["user"],
            date=data["date"],
            defaults={
                "office": data["office"],
                "shift": data["shift"],
                "note": data.get("note", ""),
            }
        )
        return Response(RosterAssignmentSerializer(obj).data, status=200)


# ============================================================
# ADMIN USERS LIST
# ============================================================
class AdminUserListView(APIView):
    permission_classes = [IsAdminUser]

    def get(self, request):
        q = (request.query_params.get("q") or "").strip()
        qs = User.objects.filter(is_staff=False, is_superuser=False).order_by("-id")
        if q:
            qs = qs.filter(Q(email__icontains=q) | Q(full_name__icontains=q))
        return Response(AdminUserListSerializer(qs, many=True).data)


# ============================================================
# ADMIN DASHBOARD / REPORT / EXPORT (your code intact)
# ============================================================
OFFICE_START = time(10, 0, 0)
OFFICE_END   = time(19, 0, 0)

def _parse_date(s: str) -> date:
    return datetime.strptime(s, "%Y-%m-%d").date()

def _date_range_list(d1: date, d2: date):
    out = []
    cur = d1
    while cur <= d2:
        out.append(cur)
        cur += timedelta(days=1)
    return out

def _parse_user_ids(s: str):
    out = []
    for part in (s or "").split(","):
        part = part.strip()
        if part.isdigit():
            out.append(int(part))
    return out

def _minutes_late(check_in_dt):
    if not check_in_dt:
        return 0
    t = localtime(check_in_dt).time()
    if t <= OFFICE_START:
        return 0
    delta = datetime.combine(date.today(), t) - datetime.combine(date.today(), OFFICE_START)
    return int(delta.total_seconds() // 60)

def _build_attendance_rows(from_date: date, to_date: date, user_ids=None, office_id=None):
    user_ids = user_ids or []

    users_qs = User.objects.filter(is_staff=False, is_superuser=False).order_by("id")
    if user_ids:
        users_qs = users_qs.filter(id__in=user_ids)

    att_qs = Attendance.objects.select_related("user", "office").filter(
        date__gte=from_date, date__lte=to_date
    )
    if user_ids:
        att_qs = att_qs.filter(user_id__in=user_ids)
    if office_id:
        att_qs = att_qs.filter(office_id=office_id)

    att_map = {}
    for a in att_qs:
        att_map[(a.user_id, a.date)] = a

    days = _date_range_list(from_date, to_date)
    rows = []
    per_user_summary = {}

    def fmt_dt(dt):
        if not dt:
            return ""
        return localtime(dt).strftime("%H:%M:%S")

    for u in users_qs:
        present = 0
        absent = 0
        late = 0

        for d in days:
            a = att_map.get((u.id, d))
            if not a:
                status = "ABSENT"
                late_min = 0
                office_name = ""
                absent += 1
                cin = ""
                cout = ""
            else:
                status = "PRESENT"
                office_name = a.office.name if a.office_id else ""
                late_min = _minutes_late(a.check_in_time)
                present += 1
                if late_min > 0:
                    late += 1
                cin = fmt_dt(a.check_in_time)
                cout = fmt_dt(a.check_out_time)

            rows.append({
                "date": str(d),
                "user_id": u.id,
                "email": u.email,
                "full_name": u.full_name or "",
                "office": office_name,
                "check_in_time": cin,
                "check_out_time": cout,
                "status": status,
                "late_minutes": late_min,
            })

        per_user_summary[u.id] = {
            "user_id": u.id,
            "email": u.email,
            "full_name": u.full_name or "",
            "total_days": len(days),
            "present_days": present,
            "absent_days": absent,
            "late_days": late,
        }

    return rows, list(per_user_summary.values())


class AdminDashboardSummaryView(APIView):
    permission_classes = [IsAdminUser]

    def get(self, request):
        days = int(request.query_params.get("days") or 30)
        to_date = localdate()
        from_date = to_date - timedelta(days=days - 1)

        rows, per_user = _build_attendance_rows(from_date, to_date, user_ids=None, office_id=None)

        total_present = sum(x["present_days"] for x in per_user)
        total_absent  = sum(x["absent_days"] for x in per_user)
        total_late    = sum(x["late_days"] for x in per_user)

        return Response({
            "from": str(from_date),
            "to": str(to_date),
            "days": days,
            "overall": {
                "total_users": len(per_user),
                "total_present_days": total_present,
                "total_absent_days": total_absent,
                "total_late_days": total_late,
            },
            "per_user": per_user,
        })


class AdminAttendanceReportView(APIView):
    permission_classes = [IsAdminUser]

    def get(self, request):
        days = request.query_params.get("days")
        from_s = request.query_params.get("from")
        to_s = request.query_params.get("to")
        user_id = request.query_params.get("user_id")
        user_ids = request.query_params.get("user_ids")
        office_id = request.query_params.get("office_id")

        office_id = int(office_id) if (office_id and office_id.isdigit()) else None

        ids = []
        if user_id and user_id.isdigit():
            ids = [int(user_id)]
        elif user_ids:
            ids = _parse_user_ids(user_ids)

        if from_s and to_s:
            from_date = _parse_date(from_s)
            to_date = _parse_date(to_s)
        else:
            d = int(days) if (days and str(days).isdigit()) else 7
            to_date = localdate()
            from_date = to_date - timedelta(days=d - 1)

        rows, summary = _build_attendance_rows(from_date, to_date, user_ids=ids, office_id=office_id)

        total_present = sum(x["present_days"] for x in summary)
        total_absent  = sum(x["absent_days"] for x in summary)
        total_late    = sum(x["late_days"] for x in summary)

        return Response({
            "from": str(from_date),
            "to": str(to_date),
            "days": (to_date - from_date).days + 1,
            "filters": {"user_ids": ids, "office_id": office_id},
            "overall": {
                "total_users": len(summary),
                "total_present_days": total_present,
                "total_absent_days": total_absent,
                "total_late_days": total_late,
            },
            "summary": summary,
            "rows": rows,
        })


class AdminAttendanceExportView(APIView):
    permission_classes = [IsAdminUser]

    def get(self, request):
        fmt = (request.query_params.get("format") or "xlsx").lower()
        days = request.query_params.get("days")
        from_s = request.query_params.get("from")
        to_s = request.query_params.get("to")
        user_id = request.query_params.get("user_id")
        user_ids = request.query_params.get("user_ids")
        office_id = request.query_params.get("office_id")

        office_id = int(office_id) if (office_id and office_id.isdigit()) else None

        ids = []
        if user_id and str(user_id).isdigit():
            ids = [int(user_id)]
        elif user_ids:
            ids = _parse_user_ids(user_ids)

        if from_s and to_s:
            from_date = _parse_date(from_s)
            to_date = _parse_date(to_s)
        else:
            d = int(days) if (days and str(days).isdigit()) else 7
            to_date = localdate()
            from_date = to_date - timedelta(days=d - 1)

        rows, summary = _build_attendance_rows(from_date, to_date, user_ids=ids, office_id=office_id)

        filename = f"attendance_{from_date}_to_{to_date}"
        if ids:
            filename += f"_users_{len(ids)}"

        if fmt == "csv":
            return self._export_csv(rows, summary, filename)
        if fmt == "xlsx":
            return self._export_xlsx(rows, summary, filename)
        if fmt == "pdf":
            return self._export_pdf(rows, summary, filename)

        return Response({"detail": "Invalid format. Use xlsx/pdf/csv."}, status=400)

    def _export_csv(self, rows, summary, filename):
        resp = HttpResponse(content_type="text/csv")
        resp["Content-Disposition"] = f'attachment; filename="{filename}.csv"'
        w = csv.writer(resp)

        w.writerow(["SUMMARY"])
        w.writerow(["user_id", "email", "name", "total_days", "present_days", "absent_days", "late_days"])
        for s in summary:
            w.writerow([s["user_id"], s["email"], s["full_name"], s["total_days"], s["present_days"], s["absent_days"], s["late_days"]])

        w.writerow([])
        w.writerow(["DETAIL"])
        w.writerow(["date", "user_id", "email", "name", "office", "check_in", "check_out", "status", "late_minutes"])
        for r in rows:
            w.writerow([r["date"], r["user_id"], r["email"], r["full_name"], r["office"], r["check_in_time"], r["check_out_time"], r["status"], r["late_minutes"]])

        return resp

    def _export_xlsx(self, rows, summary, filename):
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance"

        bold = Font(bold=True)
        red_fill = PatternFill("solid", fgColor="FFEEEE")
        red_font = Font(color="CC0000", bold=True)
        orange_fill = PatternFill("solid", fgColor="FFF3E0")
        orange_font = Font(color="E65100", bold=True)

        ws.append(["SUMMARY"])
        ws["A1"].font = bold
        ws.append(["user_id", "email", "name", "total_days", "present_days", "absent_days", "late_days"])
        for c in range(1, 8):
            ws.cell(row=2, column=c).font = bold

        for s in summary:
            ws.append([s["user_id"], s["email"], s["full_name"], s["total_days"], s["present_days"], s["absent_days"], s["late_days"]])

        ws.append([])
        start_row = ws.max_row + 1

        ws.append(["DETAIL"])
        ws.cell(row=start_row, column=1).font = bold

        ws.append(["date", "user_id", "email", "name", "office", "check_in", "check_out", "status", "late_minutes"])
        header_row = ws.max_row
        for c in range(1, 10):
            ws.cell(row=header_row, column=c).font = bold

        for r in rows:
            ws.append([r["date"], r["user_id"], r["email"], r["full_name"], r["office"], r["check_in_time"], r["check_out_time"], r["status"], r["late_minutes"]])
            rr = ws.max_row
            if r["status"] == "ABSENT":
                for c in range(1, 10):
                    ws.cell(row=rr, column=c).fill = red_fill
                ws.cell(row=rr, column=8).font = red_font
            elif r["late_minutes"] and r["late_minutes"] > 0:
                for c in range(1, 10):
                    ws.cell(row=rr, column=c).fill = orange_fill
                ws.cell(row=rr, column=9).font = orange_font

        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)

        resp = HttpResponse(
            bio.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        resp["Content-Disposition"] = f'attachment; filename="{filename}.xlsx"'
        return resp

    def _export_pdf(self, rows, summary, filename):
        buff = BytesIO()
        doc = SimpleDocTemplate(buff, pagesize=landscape(A4))
        styles = getSampleStyleSheet()

        elems = []
        elems.append(Paragraph("Attendance Report", styles["Title"]))
        elems.append(Spacer(1, 8))

        elems.append(Paragraph("Summary", styles["Heading2"]))
        sum_data = [["User", "Email", "Total", "Present", "Absent", "Late"]]
        for s in summary:
            sum_data.append([s["full_name"] or str(s["user_id"]), s["email"], s["total_days"], s["present_days"], s["absent_days"], s["late_days"]])

        sum_table = Table(sum_data, repeatRows=1)
        sum_table.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), colors.lightgrey),
            ("GRID", (0,0), (-1,-1), 0.5, colors.grey),
            ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ]))
        elems.append(sum_table)
        elems.append(Spacer(1, 12))

        elems.append(Paragraph("Detail", styles["Heading2"]))
        det_data = [["Date", "Name", "Email", "Office", "In", "Out", "Status", "Late(min)"]]
        for r in rows:
            det_data.append([r["date"], r["full_name"], r["email"], r["office"], r["check_in_time"], r["check_out_time"], r["status"], str(r["late_minutes"])])

        det_table = Table(det_data, repeatRows=1)
        ts = TableStyle([
            ("BACKGROUND", (0,0), (-1,0), colors.lightgrey),
            ("GRID", (0,0), (-1,-1), 0.25, colors.grey),
            ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE", (0,0), (-1,-1), 8),
        ])

        for i in range(1, len(det_data)):
            status_val = det_data[i][6]
            late_val = int(det_data[i][7]) if det_data[i][7].isdigit() else 0
            if status_val == "ABSENT":
                ts.add("TEXTCOLOR", (6,i), (6,i), colors.red)
                ts.add("BACKGROUND", (0,i), (-1,i), colors.whitesmoke)
            elif late_val > 0:
                ts.add("TEXTCOLOR", (7,i), (7,i), colors.orange)

        det_table.setStyle(ts)
        elems.append(det_table)

        doc.build(elems)
        pdf = buff.getvalue()
        buff.close()

        resp = HttpResponse(pdf, content_type="application/pdf")
        resp["Content-Disposition"] = f'attachment; filename="{filename}.pdf"'
        return resp


# ============================================================
# DAILY REPORTS
# ============================================================
def _parse_date_ymd(s: str) -> date:
    return datetime.strptime(s, "%Y-%m-%d").date()


class MyDailyReportListCreateView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        qs = DailyReport.objects.filter(user=request.user).order_by("-report_date", "-created_at")

        from_s = request.query_params.get("from")
        to_s = request.query_params.get("to")

        if from_s:
            qs = qs.filter(report_date__gte=from_s)
        if to_s:
            qs = qs.filter(report_date__lte=to_s)

        if not from_s and not to_s:
            qs = qs[:60]

        return Response(DailyReportSerializer(qs, many=True).data, status=200)

    def post(self, request):
        ser = DailyReportSerializer(data=request.data, context={"request": request})
        ser.is_valid(raise_exception=True)
        obj = ser.save()
        return Response(DailyReportSerializer(obj).data, status=201)


class AdminDailyReportListView(APIView):
    permission_classes = [IsAdminUser]

    def get(self, request):
        qs = DailyReport.objects.select_related("user").all()

        from_s = request.query_params.get("from")
        to_s = request.query_params.get("to")
        user_id = request.query_params.get("user_id")

        if from_s:
            qs = qs.filter(report_date__gte=from_s)
        if to_s:
            qs = qs.filter(report_date__lte=to_s)
        if user_id and str(user_id).isdigit():
            qs = qs.filter(user_id=int(user_id))

        qs = qs.order_by("report_date", "user__email", "created_at")
        return Response(DailyReportSerializer(qs, many=True).data, status=200)


class AdminDailyReportExportPDFView(APIView):
    permission_classes = [IsAdminUser]

    def get(self, request):
        from_s = request.query_params.get("from")
        to_s = request.query_params.get("to")
        user_id = request.query_params.get("user_id")

        if from_s and to_s:
            from_date = _parse_date_ymd(from_s)
            to_date = _parse_date_ymd(to_s)
        else:
            to_date = localdate()
            from_date = to_date - timedelta(days=6)

        qs = DailyReport.objects.select_related("user").filter(
            report_date__gte=from_date,
            report_date__lte=to_date
        )

        if user_id and str(user_id).isdigit():
            qs = qs.filter(user_id=int(user_id))

        qs = qs.order_by("report_date", "user__email", "created_at")

        grouped = {}
        for r in qs:
            grouped.setdefault(r.report_date, []).append(r)

        buff = BytesIO()
        doc = SimpleDocTemplate(buff, pagesize=A4, topMargin=24, bottomMargin=24, leftMargin=24, rightMargin=24)
        styles = getSampleStyleSheet()
        elems = []

        title = f"Daily Reports ({from_date} to {to_date})"
        if user_id and str(user_id).isdigit():
            title += f" | user_id={user_id}"

        elems.append(Paragraph(title, styles["Title"]))
        elems.append(Spacer(1, 10))

        if not grouped:
            elems.append(Paragraph("No reports found in selected date range.", styles["Normal"]))
            doc.build(elems)
            pdf = buff.getvalue()
            buff.close()
            resp = HttpResponse(pdf, content_type="application/pdf")
            resp["Content-Disposition"] = f'attachment; filename="daily_reports_{from_date}_to_{to_date}.pdf"'
            return resp

        for i, day in enumerate(sorted(grouped.keys())):
            elems.append(Paragraph(f"Date: {day}", styles["Heading2"]))
            elems.append(Spacer(1, 6))

            data = [["Employee", "Email", "Title", "Status", "Description"]]
            for r in grouped[day]:
                data.append([
                    (r.user.full_name or "").strip() or f"User #{r.user_id}",
                    r.user.email,
                    r.title,
                    r.status,
                    (r.description or "")
                ])

            table = Table(data, repeatRows=1, colWidths=[90, 120, 120, 60, 150])
            table.setStyle(TableStyle([
                ("BACKGROUND", (0,0), (-1,0), colors.lightgrey),
                ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
                ("GRID", (0,0), (-1,-1), 0.25, colors.grey),
                ("VALIGN", (0,0), (-1,-1), "TOP"),
                ("FONTSIZE", (0,0), (-1,-1), 8),
                ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.whitesmoke, colors.white]),
            ]))

            elems.append(table)
            if i != len(grouped.keys()) - 1:
                elems.append(PageBreak())

        doc.build(elems)
        pdf = buff.getvalue()
        buff.close()

        resp = HttpResponse(pdf, content_type="application/pdf")
        resp["Content-Disposition"] = f'attachment; filename="daily_reports_{from_date}_to_{to_date}.pdf"'
        return resp


class MyDailyReportExportPDFView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from_s = request.query_params.get("from")
        to_s = request.query_params.get("to")

        if from_s and to_s:
            from_date = _parse_date_ymd(from_s)
            to_date = _parse_date_ymd(to_s)
        else:
            to_date = localdate()
            from_date = to_date - timedelta(days=6)

        qs = DailyReport.objects.filter(
            user=request.user,
            report_date__gte=from_date,
            report_date__lte=to_date
        ).order_by("report_date", "created_at")

        grouped = {}
        for r in qs:
            grouped.setdefault(r.report_date, []).append(r)

        buff = BytesIO()
        doc = SimpleDocTemplate(buff, pagesize=A4, topMargin=24, bottomMargin=24, leftMargin=24, rightMargin=24)
        styles = getSampleStyleSheet()
        elems = []

        elems.append(Paragraph(f"My Daily Reports ({from_date} to {to_date})", styles["Title"]))
        elems.append(Spacer(1, 10))

        if not grouped:
            elems.append(Paragraph("No reports found in selected date range.", styles["Normal"]))
            doc.build(elems)
            pdf = buff.getvalue()
            buff.close()
            resp = HttpResponse(pdf, content_type="application/pdf")
            resp["Content-Disposition"] = f'attachment; filename="my_daily_reports_{from_date}_to_{to_date}.pdf"'
            return resp

        for i, day in enumerate(sorted(grouped.keys())):
            elems.append(Paragraph(f"Date: {day}", styles["Heading2"]))
            elems.append(Spacer(1, 6))

            data = [["Title", "Status", "Description"]]
            for r in grouped[day]:
                data.append([r.title, r.status, r.description or ""])

            table = Table(data, repeatRows=1, colWidths=[170, 70, 250])
            table.setStyle(TableStyle([
                ("BACKGROUND", (0,0), (-1,0), colors.lightgrey),
                ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
                ("GRID", (0,0), (-1,-1), 0.25, colors.grey),
                ("VALIGN", (0,0), (-1,-1), "TOP"),
                ("FONTSIZE", (0,0), (-1,-1), 9),
                ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.whitesmoke, colors.white]),
            ]))
            elems.append(table)

            if i != len(grouped.keys()) - 1:
                elems.append(PageBreak())

        doc.build(elems)
        pdf = buff.getvalue()
        buff.close()

        resp = HttpResponse(pdf, content_type="application/pdf")
        resp["Content-Disposition"] = f'attachment; filename="my_daily_reports_{from_date}_to_{to_date}.pdf"'
        return resp


class MyDailyReportUpdateView(APIView):
    permission_classes = [IsAuthenticated]

    def patch(self, request, report_id: int):
        obj = DailyReport.objects.filter(id=report_id, user=request.user).first()
        if not obj:
            return Response({"detail": "Not found"}, status=404)

        allowed = {}
        if "status" in request.data:
            allowed["status"] = request.data.get("status")
        if "title" in request.data:
            allowed["title"] = request.data.get("title")
        if "description" in request.data:
            allowed["description"] = request.data.get("description")

        ser = DailyReportSerializer(obj, data=allowed, partial=True, context={"request": request})
        ser.is_valid(raise_exception=True)
        ser.save()
        return Response(ser.data, status=200)
